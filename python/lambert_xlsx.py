from openpyxl import load_workbook
wb = load_workbook('0601_94_aggregate.xlsx')
print(wb.sheetnames)
print(wb)
for sheet in wb:
	print(sheet.title)
	if sheet.title == 'Dodd':
		for row in sheet.rows:
			for cell in row:
				print(str(cell.value) + ", ", end='')
			print("")
			print(len(row))
			print("Le:)